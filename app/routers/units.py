from fastapi import APIRouter, HTTPException, status, Query, Depends, Header, Response
from typing import List, Optional, Dict
import uuid
from datetime import datetime
from app.models.units import (
    UnitCalculateRequest, UnitCalculateResponse, 
    UnitEstimateRequest, UnitEstimateResponse,
    UnitType
)
from app.models.internal_counter import (
    InternalCounterRequest, InternalCounterResponse,
    InternalCounterOptions
)
from app.models.edge_band import EdgeBreakdownResponse, EdgeType
from app.services.unit_calculators import (
    calculate_unit_parts,
    calculate_total_edge_band,
    calculate_total_area,
    calculate_material_usage
)
from app.services.internal_counter_calculator import (
    calculate_internal_counter_parts,
    calculate_internal_total_edge_band,
    calculate_internal_total_area,
    calculate_internal_material_usage
)
from app.services.edge_band_calculator import (
    calculate_edge_breakdown,
    calculate_total_edge_meters,
    calculate_edge_cost
)
from app.database import get_database
from app.models.settings import SettingsModel
from app.services.auth_service import (
    TokenData, 
    get_user_by_id, 
    get_user_units_count
)
import jwt
from app.services.auth_service import SECRET_KEY, ALGORITHM
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

router = APIRouter()

# Map Arabic labels to unit types
UNIT_TYPE_LABELS = {
    "ground": "خزانة سفلية",
    "wall": "خزانة علوية",
    "double_door": "خزانة ضلفتين",
    "sink_ground": "خزانة حوض",
    "tall": "خزانة طويلة",
    "drawer": "وحدة أدراج",
    "corner": "خزانة زاوية"
}

async def get_current_user_from_token(token: str) -> TokenData:
    """Get current user from token"""
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        role: str = payload.get("role")
        if user_id is None:
            raise credentials_exception
        token_data = TokenData(user_id=user_id, role=role)
    except jwt.PyJWTError:
        raise credentials_exception
    return token_data

async def get_settings_model() -> SettingsModel:
    """Get settings model from database"""
    from app.routers.settings import get_settings_from_db
    settings_doc = await get_settings_from_db()
    settings_doc.pop("_id", None)
    
    try:
        return SettingsModel(**settings_doc)
    except Exception as validation_error:
         # If DB data is invalid/outdated, log it and return defaults
        print(f"WARNING: Settings validation failed in units router: {validation_error}. Returning defaults.")
        return SettingsModel()

@router.get("/types", response_model=List[Dict[str, str]])
async def get_unit_types():
    """
    جلب أنواع الوحدات المتاحة في النظام
    
    Returns:
    - List[Dict[str, str]]: قائمة بأنواع الوحدات مع تسمياتها
    """
    try:
        unit_types = []
        for unit_type in UnitType:
            unit_types.append({
                "value": unit_type.value,
                "label": UNIT_TYPE_LABELS.get(unit_type.value, unit_type.value)
            })
        return unit_types
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error retrieving unit types: {str(e)}"
        )

@router.post("/calculate", response_model=UnitCalculateResponse)
async def calculate_unit(request: UnitCalculateRequest, authorization: str = Header(None)):
    """
    حساب تفاصيل الوحدة (قطع الألواح وقياساتها) دون حفظ
    
    Parameters:
    - request: UnitCalculateRequest - تفاصيل الوحدة المطلوب حسابها
    - authorization: Header - توكن المستخدم (اختياري)
    
    Returns:
    - UnitCalculateResponse - تفاصيل القطع والأبعاد
    """
    try:
        # Extract token from Authorization header if provided
        if authorization and authorization.startswith("Bearer "):
            token = authorization[len("Bearer "):]
            try:
                token_data = await get_current_user_from_token(token)
                
                # Get user to check subscription limits (non-admin users only)
                if token_data.role != "admin":
                    user = await get_user_by_id(token_data.user_id)
                    if user:
                        # Check unlimited expiry
                        is_unlimited = user.subscription.is_unlimited_units
                        if is_unlimited and user.subscription.unlimited_expiry_date:
                            if datetime.utcnow() > user.subscription.unlimited_expiry_date:
                                is_unlimited = False
                        
                        if not is_unlimited:
                            # Get user's current unit count for the month
                            current_units_count = await get_user_units_count(token_data.user_id, 30)
                            
                            # Check if user has reached their limit
                            if current_units_count >= user.subscription.max_units_per_month:
                                raise HTTPException(
                                    status_code=status.HTTP_403_FORBIDDEN,
                                    detail=f"لقد بلغت الحد الأقصى من الوحدات ({user.subscription.max_units_per_month} وحدة/شهر). يرجى التواصل مع المسؤول لزيادة الحد."
                                )
            except:
                # If token is invalid, continue without user tracking
                pass
        
        # Get settings
        settings = await get_settings_model()
        
        # Calculate parts (convert cm to mm for calculation)
        parts = calculate_unit_parts(
            unit_type=request.type.value,
            width_cm=request.width_cm,
            height_cm=request.height_cm,
            depth_cm=request.depth_cm,
            shelf_count=request.shelf_count,
            door_count=request.door_count,
            door_type=request.door_type.value,
            flip_door_height=request.flip_door_height,
            bottom_door_height=request.bottom_door_height,
            oven_height=request.oven_height,
            microwave_height=request.microwave_height,
            vent_height=request.vent_height,
            width_2_cm=request.width_2_cm,
            depth_2_cm=request.depth_2_cm,
            drawer_count=request.drawer_count,
            drawer_height_cm=request.drawer_height_cm,
            fixed_part_cm=request.fixed_part_cm,
            settings=settings
        )
        
        # Calculate material usage
        total_area = calculate_total_area(parts)
        total_edge_meters = calculate_total_edge_band(parts)
        material_usage = calculate_material_usage(total_area, total_edge_meters, settings)
        
        # Convert to cm for response
        return UnitCalculateResponse(
            unit_id=str(uuid.uuid4()),
            type=request.type,
            width_cm=request.width_cm,
            height_cm=request.height_cm,
            depth_cm=request.depth_cm,
            shelf_count=request.shelf_count,
            parts=parts,
            total_edge_band_m=total_edge_meters,
            total_area_m2=total_area,
            material_usage=material_usage
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error calculating unit: {str(e)}"
        )

@router.post("/estimate", response_model=UnitEstimateResponse)
async def estimate_unit_cost(request: UnitEstimateRequest, authorization: str = Header(None)):
    """
    تقدير تكلفة الوحدة
    
    Parameters:
    - request: UnitEstimateRequest - تفاصيل الوحدة المطلوب تقدير تكلفتها
    - authorization: Header - توكن المستخدم (اختياري)
    
    Returns:
    - UnitEstimateResponse - تقدير التكلفة
    """
    try:
        # Extract token from Authorization header if provided
        user_id = None
        if authorization and authorization.startswith("Bearer "):
            token = authorization[len("Bearer "):]
            try:
                token_data = await get_current_user_from_token(token)
                user_id = token_data.user_id
                
                # Get user to check subscription limits (non-admin users only)
                if token_data.role != "admin":
                    user = await get_user_by_id(token_data.user_id)
                    if user:
                        # Check unlimited expiry
                        is_unlimited = user.subscription.is_unlimited_units
                        if is_unlimited and user.subscription.unlimited_expiry_date:
                            if datetime.utcnow() > user.subscription.unlimited_expiry_date:
                                is_unlimited = False

                        if not is_unlimited:
                            # Get user's current unit count for the month
                            current_units_count = await get_user_units_count(token_data.user_id, 30)
                            
                            # Check if user has reached their limit
                            if current_units_count >= user.subscription.max_units_per_month:
                                raise HTTPException(
                                    status_code=status.HTTP_403_FORBIDDEN,
                                    detail=f"لقد بلغت الحد الأقصى من الوحدات ({user.subscription.max_units_per_month} وحدة/شهر). يرجى التواصل مع المسؤول لزيادة الحد."
                                )
            except:
                # If token is invalid, continue without user tracking
                pass
        
        # Get settings
        settings = await get_settings_model()
        
        # Calculate parts (convert cm to mm for calculation)
        parts = calculate_unit_parts(
            unit_type=request.type.value,
            width_cm=request.width_cm,
            height_cm=request.height_cm,
            depth_cm=request.depth_cm,
            shelf_count=request.shelf_count,
            door_count=request.door_count,
            door_type=request.door_type.value,
            flip_door_height=request.flip_door_height,
            bottom_door_height=request.bottom_door_height,
            oven_height=request.oven_height,
            microwave_height=request.microwave_height,
            vent_height=request.vent_height,
            width_2_cm=request.width_2_cm,
            depth_2_cm=request.depth_2_cm,
            drawer_count=request.drawer_count,
            drawer_height_cm=request.drawer_height_cm,
            fixed_part_cm=request.fixed_part_cm,
            settings=settings
        )
        
        # Calculate total area
        total_area = calculate_total_area(parts)
        
        # Calculate edge band
        total_edge_meters = calculate_total_edge_band(parts)
        
        # Calculate material usage
        material_usage = calculate_material_usage(total_area, total_edge_meters, settings)
        
        # Calculate cost based on materials
        total_cost = 0.0
        plywood_cost = 0.0
        edge_band_cost = 0.0
        
        if settings.materials.get("plywood_sheet"):
            plywood = settings.materials["plywood_sheet"]
            if plywood.price_per_sheet and material_usage.get("ألواح الخشب"):
                plywood_cost = material_usage["ألواح الخشب"] * plywood.price_per_sheet
                total_cost += plywood_cost
        
        # Calculate edge band cost
        if total_edge_meters and settings.materials.get("edge_band_per_meter"):
            edge_band = settings.materials["edge_band_per_meter"]
            if edge_band.price_per_meter and material_usage.get("شريط الحافة"):
                edge_band_cost = material_usage["شريط الحافة"] * edge_band.price_per_meter
                total_cost += edge_band_cost
        
        # Convert to cm for response
        return UnitEstimateResponse(
            unit_id=str(uuid.uuid4()),
            type=request.type,
            width_cm=request.width_cm,
            height_cm=request.height_cm,
            depth_cm=request.depth_cm,
            shelf_count=request.shelf_count,
            parts=parts,
            total_edge_band_m=total_edge_meters,
            total_area_m2=total_area,
            material_usage=material_usage,
            cost_breakdown={
                "ألواح الخشب": plywood_cost,
                "شريط الحافة": edge_band_cost
            },
            total_cost=total_cost
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error estimating unit cost: {str(e)}"
        )

@router.get("/{unit_id}", response_model=UnitCalculateResponse)
async def get_unit(unit_id: str):
    """
    جلب تفاصيل وحدة محفوظة
    
    Parameters:
    - unit_id: str - معرف الوحدة
    
    Returns:
    - UnitCalculateResponse - تفاصيل الوحدة
    """
    try:
        db = get_database()
        if db is None:
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="Database connection not available"
            )
        
        units_collection = db.units
        unit_doc = await units_collection.find_one({"_id": unit_id})
        
        if unit_doc is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Unit not found"
            )
        
        # Get settings for cost calculation
        settings = await get_settings_model()
        
        # Convert database document to response format
        response_data = unit_doc.copy()
        
        # Map database field names to response field names
        if "_id" in response_data:
            response_data["unit_id"] = str(response_data["_id"])
            del response_data["_id"]
        
        if "parts_calculated" in response_data:
            # Convert parts_calculated to Part objects
            from app.models.units import Part
            parts = [Part(**part_data) for part_data in response_data["parts_calculated"]]
            response_data["parts"] = parts
            del response_data["parts_calculated"]
        
        if "edge_band_m" in response_data:
            response_data["total_edge_band_m"] = response_data["edge_band_m"]
            del response_data["edge_band_m"]
        
        # Recalculate costs from material usage if not present or zero
        if "material_usage" in response_data:
            material_usage = response_data["material_usage"]
            total_edge_meters = response_data.get("total_edge_band_m", 0)
            
            # Calculate cost based on materials
            total_cost = 0.0
            cost_breakdown = {}
            
            # Calculate plywood cost - using the correct material key from settings
            if settings.materials.get("plywood_sheet"):
                plywood = settings.materials["plywood_sheet"]
                if plywood.price_per_sheet and material_usage.get("ألواح الخشب"):
                    plywood_cost = material_usage["ألواح الخشب"] * plywood.price_per_sheet
                    cost_breakdown["ألواح الخشب"] = plywood_cost
                    total_cost += plywood_cost
            
            # Calculate edge band cost - using the correct material key from settings
            if settings.materials.get("edge_band_per_meter"):
                edge_band = settings.materials["edge_band_per_meter"]
                if edge_band.price_per_meter and material_usage.get("شريط الحافة"):
                    edge_band_cost = material_usage["شريط الحافة"] * edge_band.price_per_meter
                    cost_breakdown["شريط الحافة"] = edge_band_cost
                    total_cost += edge_band_cost
            
            response_data["total_cost"] = total_cost
            response_data["cost_breakdown"] = cost_breakdown
        elif "total_cost" not in response_data:
            response_data["total_cost"] = 0.0
            
        if "cost_breakdown" not in response_data:
            response_data["cost_breakdown"] = {}
        
        return UnitCalculateResponse(**response_data)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error retrieving unit: {str(e)}"
        )

@router.post("", response_model=UnitCalculateResponse)
async def save_unit(request: UnitCalculateRequest, authorization: str = Header(None)):
    """
    حفظ وحدة في قاعدة البيانات
    
    Parameters:
    - request: UnitCalculateRequest - تفاصيل الوحدة المطلوب حفظها
    - authorization: Header - توكن المستخدم
    
    Returns:
    - UnitCalculateResponse - تفاصيل الوحدة المحفوظة
    """
    try:
        # Extract token from Authorization header
        if not authorization or not authorization.startswith("Bearer "):
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid authorization header"
            )
        
        token = authorization[len("Bearer "):]
        token_data = await get_current_user_from_token(token)
        
        # Get settings
        settings = await get_settings_model()
        
        # Calculate parts (convert cm to mm for calculation)
        parts = calculate_unit_parts(
            unit_type=request.type.value,
            width_cm=request.width_cm,
            height_cm=request.height_cm,
            depth_cm=request.depth_cm,
            shelf_count=request.shelf_count,
            door_count=request.door_count,
            door_type=request.door_type.value,
            flip_door_height=request.flip_door_height,
            bottom_door_height=request.bottom_door_height,
            oven_height=request.oven_height,
            microwave_height=request.microwave_height,
            vent_height=request.vent_height,
            width_2_cm=request.width_2_cm,
            depth_2_cm=request.depth_2_cm,
            drawer_count=request.drawer_count,
            drawer_height_cm=request.drawer_height_cm,
            fixed_part_cm=request.fixed_part_cm,
            settings=settings
        )
        
        # Calculate total area
        total_area = calculate_total_area(parts)
        
        # Calculate edge band
        total_edge_meters = calculate_total_edge_band(parts)
        
        # Calculate material usage
        material_usage = calculate_material_usage(total_area, total_edge_meters, settings)
        
        # Calculate cost based on materials
        total_cost = 0.0
        plywood_cost = 0.0
        edge_band_cost = 0.0
        
        if settings.materials.get("plywood_sheet"):
            plywood = settings.materials["plywood_sheet"]
            if plywood.price_per_sheet and material_usage.get("ألواح الخشب"):
                plywood_cost = material_usage["ألواح الخشب"] * plywood.price_per_sheet
                total_cost += plywood_cost
        
        # Calculate edge band cost
        if total_edge_meters and settings.materials.get("edge_band_per_meter"):
            edge_band = settings.materials["edge_band_per_meter"]
            if edge_band.price_per_meter and material_usage.get("شريط الحافة"):
                edge_band_cost = material_usage["شريط الحافة"] * edge_band.price_per_meter
                total_cost += edge_band_cost
        
        # Create unit document (store in cm)
        unit_id = str(uuid.uuid4())
        unit_doc = {
            "_id": unit_id,
            "type": request.type,
            "width_cm": request.width_cm,
            "height_cm": request.height_cm,
            "depth_cm": request.depth_cm,
            "shelf_count": request.shelf_count,
            "parts_calculated": [part.model_dump() for part in parts],
            "edge_band_m": total_edge_meters,
            "total_area_m2": total_area,
            "material_usage": material_usage,
            "price_estimate": total_cost,
            "cost_breakdown": {
                "ألواح الخشب": plywood_cost,
                "شريط الحافة": edge_band_cost
            },
            "created_by": token_data.user_id,  # Track who created the unit
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow()
        }
        
        # Save to database
        db = get_database()
        if db is None:
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="Database connection not available"
            )
        
        units_collection = db.units
        await units_collection.insert_one(unit_doc)
        
        # Return response (convert to cm for response)
        response_data = unit_doc.copy()
        response_data["unit_id"] = response_data.pop("_id")
        # Add the field that UnitCalculateResponse expects
        response_data["total_edge_band_m"] = response_data.pop("edge_band_m")
        response_data["parts"] = parts
        # Ensure cost fields are present
        if "price_estimate" in response_data:
            response_data["total_cost"] = response_data["price_estimate"]
        else:
            response_data["total_cost"] = 0.0
            
        if "cost_breakdown" not in response_data:
            response_data["cost_breakdown"] = {}
        return UnitCalculateResponse(**response_data)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error saving unit: {str(e)}"
        )

@router.post("/{unit_id}/internal-counter/calculate", response_model=InternalCounterResponse)
async def calculate_internal_counter(unit_id: str, request: InternalCounterRequest):
    """
    حساب القطع الداخلية للكونتر
    
    يحسب القطع الداخلية مثل:
    - القاعدة الداخلية
    - المرآة الأمامية
    - الرف الداخلي
    - الأدرج (قاع، جوانب، خلفية)
    
    بناءً على أبعاد الوحدة المحفوظة وخيارات القطع الداخلية
    """
    try:
        db = get_database()
        if db is None:
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="Database connection not available"
            )
        
        # جلب بيانات الوحدة
        unit_doc = await db.units.find_one({"_id": unit_id})
        
        if unit_doc is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Unit with id {unit_id} not found"
            )
        
        # جلب الإعدادات
        settings = await get_settings_model()
        
        # استخراج بيانات الوحدة
        unit_type = UnitType(unit_doc["type"])
        unit_width_cm = unit_doc["width_cm"]
        unit_height_cm = unit_doc["height_cm"]
        unit_depth_cm = unit_doc["depth_cm"]
        
        # استخدام الخيارات من الطلب أو القيم الافتراضية
        options = request.options if request.options else InternalCounterOptions()
        
        # حساب القطع الداخلية
        internal_parts = calculate_internal_counter_parts(
            unit_type=unit_type,
            unit_width_cm=unit_width_cm,
            unit_height_cm=unit_height_cm,
            unit_depth_cm=unit_depth_cm,
            settings=settings,
            options=options
        )
        
        # حساب الإجماليات
        total_edge_band_m = calculate_internal_total_edge_band(internal_parts)
        total_area_m2 = calculate_internal_total_area(internal_parts)
        
        # حساب استخدام المواد (يستخدم settings مباشرة)
        material_usage = calculate_internal_material_usage(
            total_area_m2=total_area_m2,
            edge_band_m=total_edge_band_m,
            settings=settings
        )
        
        # حفظ القطع الداخلية في الوحدة (update)
        await db.units.update_one(
            {"_id": unit_id},
            {
                "$set": {
                    "internal_counter_parts": [part.model_dump() for part in internal_parts],
                    "internal_counter_edge_band_m": total_edge_band_m,
                    "internal_counter_total_area_m2": total_area_m2,
                    "internal_counter_material_usage": material_usage,
                    "updated_at": datetime.utcnow()
                }
            }
        )
        
        return InternalCounterResponse(
            unit_id=unit_id,
            unit_type=unit_type,
            parts=internal_parts,
            total_edge_band_m=round(total_edge_band_m, 2),
            total_area_m2=round(total_area_m2, 4),
            material_usage=material_usage
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error calculating internal counter: {str(e)}"
        )

@router.get("/{unit_id}/edge-breakdown", response_model=EdgeBreakdownResponse)
async def get_edge_breakdown(unit_id: str, edge_type: Optional[str] = None):
    """
    جلب تفاصيل توزيع الشريط للوحدة
    
    يعرض تفاصيل توزيع الشريط لكل قطعة في الوحدة:
    - الحواف التي تحتاج شريط (top, bottom, left, right)
    - طول كل حافة بالمليمتر والمتر
    - نوع الشريط (خشبي/PVC)
    - إجمالي متر الشريط لكل قطعة
    - التكلفة الإجمالية
    
    Parameters:
    - unit_id: معرف الوحدة
    - edge_type: نوع الشريط (wood/pvc) - اختياري
    """
    try:
        db = get_database()
        if db is None:
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="Database connection not available"
            )
        
        # جلب بيانات الوحدة
        unit_doc = await db.units.find_one({"_id": unit_id})
        
        if unit_doc is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Unit with id {unit_id} not found"
            )
        
        # جلب الإعدادات
        settings = await get_settings_model()
        
        # تحديد نوع الشريط
        selected_edge_type = EdgeType.PVC  # Default
        if edge_type:
            try:
                selected_edge_type = EdgeType(edge_type.lower())
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Invalid edge_type: {edge_type}. Must be 'wood' or 'pvc'"
                )
        
        # تحويل parts من MongoDB إلى Part objects
        from app.models.units import Part
        
        parts_data = unit_doc.get("parts_calculated", [])
        parts = [Part(**part_data) for part_data in parts_data]
        
        # حساب توزيع الشريط
        edge_breakdown = calculate_edge_breakdown(parts, settings, selected_edge_type)
        
        # حساب الإجماليات
        total_edge_m = calculate_total_edge_meters(edge_breakdown)
        
        # حساب التكلفة
        cost_info = calculate_edge_cost(edge_breakdown, settings)
        
        return EdgeBreakdownResponse(
            unit_id=unit_id,
            parts=edge_breakdown,
            total_edge_m=round(total_edge_m, 3),
            total_cost=cost_info["total"] if cost_info["total"] > 0 else None,
            cost_breakdown=cost_info["breakdown"] if cost_info["breakdown"] else None
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error calculating edge breakdown: {str(e)}"
        )

@router.get("/{unit_id}/export-excel", response_class=Response)
async def export_unit_to_excel(unit_id: str, authorization: str = Header(None)):
    """
    تصدير تفاصيل الوحدة إلى ملف Excel
    
    Parameters:
    - unit_id: str - معرف الوحدة
    - authorization: Header - توكن المستخدم
    
    Returns:
    - Excel file with unit parts details
    """
    try:
        # Extract token from Authorization header
        if authorization and authorization.startswith("Bearer "):
            token = authorization[len("Bearer "):]
            try:
                token_data = await get_current_user_from_token(token)
            except:
                # If token is invalid, continue without user tracking
                pass
        
        db = get_database()
        if db is None:
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="Database connection not available"
            )
        
        units_collection = db.units
        unit_doc = await units_collection.find_one({"_id": unit_id})
        
        if unit_doc is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Unit not found"
            )
        
        # Get settings for cost calculation
        settings = await get_settings_model()
        
        # Convert database document to response format
        response_data = unit_doc.copy()
        
        # Map database field names to response field names
        if "_id" in response_data:
            response_data["unit_id"] = str(response_data["_id"])
            del response_data["_id"]
        
        if "parts_calculated" in response_data:
            # Convert parts_calculated to Part objects
            from app.models.units import Part
            parts = [Part(**part_data) for part_data in response_data["parts_calculated"]]
        else:
            parts = []
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "تفاصيل الوحدة"
        
        # Set column headers with styling
        headers = ["اسم القطعة", "العرض (سم)", "الارتفاع (سم)", "الكمية", "المساحة (م²)", "طول الحافة (م)"]
        ws.append(headers)
        
        # Style the header row
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        total_qty = 0
        total_area = 0.0
        total_edge = 0.0
        
        for part in parts:
            row = [
                part.name,
                part.width_cm,
                part.height_cm,
                part.qty,
                round(part.area_m2, 2) if part.area_m2 else 0,
                round(part.edge_band_m, 2) if part.edge_band_m else 0
            ]
            ws.append(row)
            
            # Update totals
            total_qty += part.qty
            total_area += part.area_m2 or 0
            total_edge += part.edge_band_m or 0
        
        # Add totals row
        totals_row = ["المجموع", "", "", total_qty, round(total_area, 2), round(total_edge, 2)]
        ws.append(totals_row)
        
        # Style the totals row
        totals_font = Font(bold=True)
        for col in range(1, len(totals_row) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = totals_font
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        from io import BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Return Excel file as response
        headers = {
            "Content-Disposition": f"attachment; filename=unit_details_{unit_id}.xlsx",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        
        return Response(content=excel_buffer.getvalue(), headers=headers)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error exporting unit to Excel: {str(e)}"
        )
